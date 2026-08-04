import sqlite3
import pandas as pd
from typing import List, Dict, Any, Optional
from database.schema import get_connection

class ReporteController:
    """
    Controlador para la consulta y exportación de reportes detallados a Excel.
    """

    @staticmethod
    def obtener_datos_reporte(semana_id: int, refugio_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Ejecuta la consulta de solicitudes de forma agrupada por producto pedido por la familia
        (evitando el producto cartesiano), calculando a nivel de familia los integrantes totales,
        desglose por sexo y rango de edad en una sola columna.
        """
        conn = get_connection()
        cursor = conn.cursor()

        # Consulta SQL con LEFT JOIN a una subconsulta agrupada de integrantes para obtener demográficos sin duplicar filas
        query = """
        SELECT
            s.nombre_semana AS "Semana",
            r.nombre AS "Refugio",
            f.codigo_numero AS "Código Familia",
            f.nombre_representativo AS "Familia",
            COALESCE(p.nombre, ds.nombre_producto_manual) AS "Producto Solicitado",
            ds.cantidad_solicitada AS "Cantidad",
            ds.unidad_medida_solicitada AS "Unidad",
            CASE
                WHEN ds.en_inventario = 1 THEN 'Sí'
                ELSE 'No'
            END AS "Disponibilidad en Inventario (Sí / No)",
            f.id AS "FamiliaID",
            COALESCE(demog.total_integrantes, 0) AS "Total Integrantes",
            COALESCE(demog.males, 0) AS "Males",
            COALESCE(demog.females, 0) AS "Females",
            COALESCE(demog.ninos, 0) AS "Ninos",
            COALESCE(demog.adultos, 0) AS "Adultos",
            COALESCE(demog.adultos_mayores, 0) AS "AdultosMayores"
        FROM solicitudes sol
        INNER JOIN semanas s ON sol.semana_id = s.id
        INNER JOIN familias f ON sol.familia_id = f.id
        INNER JOIN refugios r ON f.refugio_id = r.id
        INNER JOIN detalles_solicitud ds ON ds.solicitud_id = sol.id
        LEFT JOIN productos p ON ds.producto_id = p.id
        LEFT JOIN (
            SELECT
                familia_id,
                COUNT(id) AS total_integrantes,
                SUM(CASE WHEN sexo = 'M' THEN 1 ELSE 0 END) AS males,
                SUM(CASE WHEN sexo = 'F' THEN 1 ELSE 0 END) AS females,
                SUM(CASE WHEN edad < 12 THEN 1 ELSE 0 END) AS ninos,
                SUM(CASE WHEN edad >= 12 AND edad <= 59 THEN 1 ELSE 0 END) AS adultos,
                SUM(CASE WHEN edad >= 60 THEN 1 ELSE 0 END) AS adultos_mayores
            FROM integrantes
            GROUP BY familia_id
        ) demog ON demog.familia_id = f.id
        WHERE s.id = ?
        """

        params = [semana_id]
        if refugio_id is not None:
            query += " AND r.id = ?"
            params.append(refugio_id)

        query += " ORDER BY f.codigo_numero"

        try:
            cursor.execute(query, params)
            rows = cursor.fetchall()

            reporte = []

            for row in rows:
                (semana, refugio, cod_familia, familia, prod_solicitado, cantidad, unidad,
                 en_inventario, familia_id, total_integrantes, males, females, ninos, adultos, adultos_mayores) = row

                # Formatear el Resumen Demográfico
                sex_str = f"{males}M / {females}F"
                age_parts = []
                if ninos > 0:
                    age_parts.append(f"{ninos} Niño" if ninos == 1 else f"{ninos} Niños")
                if adultos > 0:
                    age_parts.append(f"{adultos} Adulto" if adultos == 1 else f"{adultos} Adultos")
                if adultos_mayores > 0:
                    age_parts.append(f"{adultos_mayores} Adulto Mayor" if adultos_mayores == 1 else f"{adultos_mayores} Adultos Mayores")

                if age_parts:
                    resumen_demografico = f"{sex_str} ({', '.join(age_parts)})"
                else:
                    resumen_demografico = sex_str

                reporte.append({
                    "Semana": semana or "",
                    "Refugio": refugio or "",
                    "Código Familia": cod_familia or "",
                    "Familia": familia or "",
                    "Total Integrantes": total_integrantes,
                    "Resumen Demográfico": resumen_demografico,
                    "Producto Solicitado": prod_solicitado or "",
                    "Cantidad": cantidad if cantidad is not None else 0.0,
                    "Unidad": unidad or "",
                    "Disponibilidad en Inventario (Sí / No)": en_inventario or "No"
                })

            # Ordenar primero por Código Familia, y luego todos los "Sí" antes que "No" en Disponibilidad
            reporte.sort(key=lambda x: (x["Código Familia"], 0 if x["Disponibilidad en Inventario (Sí / No)"] == "Sí" else 1))

            return reporte
        except sqlite3.Error as e:
            print(f"Error al obtener datos de reporte: {e}")
            raise
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def exportar_a_excel(semana_id: int, refugio_id: Optional[int], filepath: str) -> None:
        """
        Genera el reporte a partir de la consulta SQL y lo exporta a un archivo Excel (.xlsx)
        con formato limpio, encabezados resaltados y auto-ajuste de columnas utilizando pandas y openpyxl.
        """
        # 1. Obtener datos
        datos = ReporteController.obtener_datos_reporte(semana_id, refugio_id)

        # 2. Crear DataFrame de pandas
        columnas = [
            "Semana", "Refugio", "Código Familia", "Familia", "Total Integrantes",
            "Resumen Demográfico", "Producto Solicitado", "Cantidad", "Unidad", "Disponibilidad en Inventario (Sí / No)"
        ]

        df = pd.DataFrame(datos, columns=columnas)

        # 3. Exportar usando pandas + openpyxl
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Reporte de Solicitudes")

            # Formateo con openpyxl
            workbook = writer.book
            worksheet = writer.sheets["Reporte de Solicitudes"]

            # Estilos de openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            header_fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid") # Turquesa oscuro del tema
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Borde fino gris
            thin_border_side = Side(border_style="thin", color="CCCCCC")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

            # Aplicar estilos a cabecera (Fila 1)
            for col_idx in range(1, len(columnas) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Aplicar estilos a las celdas de datos
            data_font = Font(name="Arial", size=10)
            data_alignment_center = Alignment(horizontal="center", vertical="center")
            data_alignment_left = Alignment(horizontal="left", vertical="center")

            # Fills y Fonts para Disponibilidad
            si_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            si_font = Font(name="Arial", size=10, color="155724")

            no_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            no_font = Font(name="Arial", size=10, color="721C24")

            # Centrar ciertas columnas como Semana, Código, Total Integrantes, Unidad, Disponibilidad
            center_columns = {1, 3, 5, 9, 10}

            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(columnas) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border

                    if col_idx in center_columns:
                        cell.alignment = data_alignment_center
                    else:
                        cell.alignment = data_alignment_left

                    # Aplicar estilos de color específicos para la columna Disponibilidad
                    if col_idx == 10:
                        if cell.value == "Sí":
                            cell.fill = si_fill
                            cell.font = si_font
                        elif cell.value == "No":
                            cell.fill = no_fill
                            cell.font = no_font

            # Ajuste automático del ancho de las columnas
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
