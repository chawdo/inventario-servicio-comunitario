import sqlite3
import pandas as pd
from typing import List, Dict, Any, Optional
from database.schema import get_connection

class InventarioController:
    """
    Controlador para gestionar la lógica de negocio y operaciones
    en la base de datos relacionadas con categorías y productos del inventario.
    """

    # ==========================================
    # GESTIÓN DE CATEGORÍAS
    # ==========================================

    @staticmethod
    def obtener_todas_categorias() -> List[Dict[str, Any]]:
        """
        Retorna todas las categorías ordenadas alfabéticamente por nombre.
        """
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT id, nombre, descripcion
                FROM categorias
                ORDER BY nombre ASC
            """)
            rows = cursor.fetchall()
            categorias = []
            for row in rows:
                categorias.append({
                    "id": row[0],
                    "nombre": row[1],
                    "descripcion": row[2] or ""
                })
            return categorias
        except sqlite3.Error as e:
            print(f"Error al obtener categorías: {e}")
            raise
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def exportar_a_excel(filepath: str) -> None:
        """
        Exporta el inventario global a un archivo Excel (.xlsx) con un formato
        limpio, encabezados estilizados y auto-ajuste de columnas.
        """
        # 1. Obtener todos los productos
        productos = InventarioController.obtener_todos_productos()

        # 2. Definir las columnas y el orden para el Excel
        columnas = [
            "ID", "Categoría", "Producto", "Presentación",
            "Tamaño/Peso Unidad", "Unidad", "Stock Disponible",
            "Precio Unidad", "Precio Kg/L"
        ]

        # 3. Transformar los diccionarios a una estructura adecuada para DataFrame
        datos = []
        for p in productos:
            datos.append({
                "ID": p["id"],
                "Categoría": p["categoria_nombre"],
                "Producto": p["nombre"],
                "Presentación": p["empaque_unidad"],
                "Tamaño/Peso Unidad": p["tamano_unidad_peso"],
                "Unidad": p["unidad_medida"],
                "Stock Disponible": p["stock_unidades"],
                "Precio Unidad": p["precio_unidad"],
                "Precio Kg/L": p["precio_kilo_litro"]
            })

        df = pd.DataFrame(datos, columns=columnas)

        # 4. Escribir usando pandas y openpyxl
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Inventario")

            workbook = writer.book
            worksheet = writer.sheets["Inventario"]

            # Estilos de openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            header_fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid") # Turquesa oscuro
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_border_side = Side(border_style="thin", color="CCCCCC")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

            # Aplicar estilos a la cabecera (Fila 1)
            for col_idx in range(1, len(columnas) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Aplicar estilos a las celdas de datos
            data_font = Font(name="Arial", size=10)
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            # Mapear alineaciones por columna
            # ID: center, Categoría: left, Producto: left, Presentación: left,
            # Tamaño: right, Unidad: center, Stock: right, Precio U: right, Precio Kg/L: right
            alignments = {
                1: align_center,  # ID
                2: align_left,    # Categoría
                3: align_left,    # Producto
                4: align_left,    # Presentación
                5: align_right,   # Tamaño/Peso Unidad
                6: align_center,  # Unidad
                7: align_right,   # Stock Disponible
                8: align_right,   # Precio Unidad
                9: align_right    # Precio Kg/L
            }

            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(columnas) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = alignments.get(col_idx, align_left)

                    # Formateo de números si corresponde
                    if col_idx in {5, 7, 8, 9}:
                        cell.number_format = "0.00"

            # Ajuste automático del ancho de las columnas
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        # Si es flotante/número, formatearlo para el cálculo de longitud
                        if isinstance(cell.value, float):
                            val_str = f"{cell.value:.2f}"
                        else:
                            val_str = str(cell.value)
                        max_len = max(max_len, len(val_str))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    @staticmethod
    def crear_categoria(nombre: str, descripcion: str) -> int:
        """
        Crea una nueva categoría de productos.
        """
        if not nombre.strip():
            raise ValueError("El nombre de la categoría no puede estar vacío.")

        conn = get_connection()
        cursor = conn.cursor()
        try:
            # Validar unicidad de nombre de categoría
            cursor.execute("SELECT id FROM categorias WHERE nombre = ?", (nombre.strip(),))
            if cursor.fetchone():
                raise ValueError(f"Ya existe una categoría con el nombre '{nombre.strip()}'.")

            cursor.execute("""
                INSERT INTO categorias (nombre, descripcion)
                VALUES (?, ?)
            """, (nombre.strip(), descripcion.strip() or None))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.Error as e:
            conn.rollback()
            print(f"Error al registrar la categoría: {e}")
            raise
        finally:
            cursor.close()
            conn.close()

    # ==========================================
    # GESTIÓN DE PRODUCTOS
    # ==========================================

    @staticmethod
    def obtener_todos_productos() -> List[Dict[str, Any]]:
        """
        Retorna todos los productos del inventario global con el nombre de su categoría.
        """
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT p.id, p.categoria_id, c.nombre AS categoria_nombre, p.nombre,
                       p.empaque_unidad, p.tamano_unidad_peso, p.unidad_medida,
                       p.stock_unidades, p.precio_unidad, p.precio_kilo_litro
                FROM productos p
                INNER JOIN categorias c ON p.categoria_id = c.id
                ORDER BY c.nombre ASC, p.nombre ASC
            """)
            rows = cursor.fetchall()
            productos = []
            for row in rows:
                productos.append({
                    "id": row[0],
                    "categoria_id": row[1],
                    "categoria_nombre": row[2],
                    "nombre": row[3],
                    "empaque_unidad": row[4],
                    "tamano_unidad_peso": row[5],
                    "unidad_medida": row[6],
                    "stock_unidades": row[7],
                    "precio_unidad": row[8],
                    "precio_kilo_litro": row[9]
                })
            return productos
        except sqlite3.Error as e:
            print(f"Error al obtener productos: {e}")
            raise
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def crear_producto(categoria_id: int, nombre: str, empaque_unidad: str,
                        tamano_unidad_peso: float, unidad_medida: str,
                        stock_unidades: float, precio_unidad: float) -> int:
        """
        Registra un nuevo producto en el inventario global.
        Calcula automáticamente precio_kilo_litro dividiendo precio_unidad entre tamano_unidad_peso.
        """
        # Validaciones de negocio
        if not categoria_id:
            raise ValueError("Debe seleccionar una categoría válida.")
        if not nombre.strip():
            raise ValueError("El nombre del producto no puede estar vacío.")
        if not empaque_unidad.strip():
            raise ValueError("La presentación/empaque del producto no puede estar vacía.")
        if tamano_unidad_peso <= 0:
            raise ValueError("El tamaño o peso por unidad debe ser un número mayor que cero.")
        if not unidad_medida.strip():
            raise ValueError("La unidad de medida no puede estar vacía.")
        if stock_unidades < 0:
            raise ValueError("El stock de unidades disponibles no puede ser negativo.")
        if precio_unidad < 0:
            raise ValueError("El precio por unidad no puede ser negativo.")

        # Cálculo automático de precio_kilo_litro
        precio_kilo_litro = precio_unidad / tamano_unidad_peso

        conn = get_connection()
        cursor = conn.cursor()
        try:
            # Validar que la categoría exista
            cursor.execute("SELECT id FROM categorias WHERE id = ?", (categoria_id,))
            if not cursor.fetchone():
                raise ValueError("La categoría seleccionada no existe.")

            cursor.execute("""
                INSERT INTO productos (categoria_id, nombre, empaque_unidad, tamano_unidad_peso,
                                       unidad_medida, stock_unidades, precio_unidad, precio_kilo_litro)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (categoria_id, nombre.strip(), empaque_unidad.strip(), tamano_unidad_peso,
                  unidad_medida.strip(), stock_unidades, precio_unidad, precio_kilo_litro))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.Error as e:
            conn.rollback()
            print(f"Error al registrar el producto: {e}")
            raise
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def actualizar_producto(producto_id: int, categoria_id: int, nombre: str, empaque_unidad: str,
                             tamano_unidad_peso: float, unidad_medida: str,
                             stock_unidades: float, precio_unidad: float) -> None:
        """
        Actualiza un producto existente en el inventario.
        Calcula automáticamente precio_kilo_litro.
        """
        # Validaciones de negocio
        if not producto_id:
            raise ValueError("ID de producto inválido.")
        if not categoria_id:
            raise ValueError("Debe seleccionar una categoría válida.")
        if not nombre.strip():
            raise ValueError("El nombre del producto no puede estar vacío.")
        if not empaque_unidad.strip():
            raise ValueError("La presentación/empaque del producto no puede estar vacía.")
        if tamano_unidad_peso <= 0:
            raise ValueError("El tamaño o peso por unidad debe ser un número mayor que cero.")
        if not unidad_medida.strip():
            raise ValueError("La unidad de medida no puede estar vacía.")
        if stock_unidades < 0:
            raise ValueError("El stock de unidades disponibles no puede ser negativo.")
        if precio_unidad < 0:
            raise ValueError("El precio por unidad no puede ser negativo.")

        # Cálculo automático de precio_kilo_litro
        precio_kilo_litro = precio_unidad / tamano_unidad_peso

        conn = get_connection()
        cursor = conn.cursor()
        try:
            # Validar que el producto exista
            cursor.execute("SELECT id FROM productos WHERE id = ?", (producto_id,))
            if not cursor.fetchone():
                raise ValueError("El producto a actualizar no existe.")

            # Validar que la categoría exista
            cursor.execute("SELECT id FROM categorias WHERE id = ?", (categoria_id,))
            if not cursor.fetchone():
                raise ValueError("La categoría seleccionada no existe.")

            cursor.execute("""
                UPDATE productos
                SET categoria_id = ?, nombre = ?, empaque_unidad = ?, tamano_unidad_peso = ?,
                    unidad_medida = ?, stock_unidades = ?, precio_unidad = ?, precio_kilo_litro = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (categoria_id, nombre.strip(), empaque_unidad.strip(), tamano_unidad_peso,
                  unidad_medida.strip(), stock_unidades, precio_unidad, precio_kilo_litro, producto_id))
            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            print(f"Error al actualizar el producto: {e}")
            raise
        finally:
            cursor.close()
            conn.close()
